"""Gera o XLSX de perfil CIMO v3.3 compatível com o profile engine do SAEC.

Fidelidade: cada campo, enum, regra e instrução foi extraído diretamente do
documento "Guia de Extração v3.3" (data/Extraction/Guia_Extracao_v3_3.md/docx).

Uso:
    python scripts/generate_cimo_profile_xlsx.py

Saída:
    data/Extraction/SAEC_Profile_CIMO_v3_3.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "Extraction" / "SAEC_Profile_CIMO_v3_3.xlsx"

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _style_headers(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


# ─────────────────────────────────────────────────────────────── meta
def _build_meta(wb: Workbook) -> None:
    ws = wb.create_sheet("meta")
    ws.append(["key", "value"])
    rows = [
        ("schema_version", "1.0"),
        ("profile_id", "cimo_rsl_ia_scm_og"),
        ("version", "3.3.0"),
        ("name", "RSL IA + SCM/SCRM em Oil & Gas (CIMO v3.3)"),
        ("framework", "CIMO"),
        ("language", "pt-BR"),
        ("domain", "Oil & Gas Supply Chain Management / Risk Management"),
        ("description",
         "Perfil de extração CIMO v3.3 para Revisão Sistemática de Literatura "
         "sobre Inteligência Artificial aplicada a Supply Chain Management (SCM) "
         "e Supply Chain Risk Management (SCRM) no setor de Oil & Gas. "
         "Fonte: Guia_Extracao_v3_3."),
        ("min_topics", "1"),
        ("max_topics", "50"),
        ("weighting_mode", "optional"),
        ("default_weight", "1.0"),
        ("output_format", "yaml"),
        ("enforce_field_names", "true"),
        ("include_sections", "true"),
        ("return_only_output", "true"),
        ("include_self_review", "true"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    _style_headers(ws, 2)


# ─────────────────────────────────────────────────────────────── fields
def _build_fields(wb: Workbook) -> None:
    ws = wb.create_sheet("fields")
    headers = [
        "id", "label", "section", "type", "required", "multiple",
        "min_length", "max_length", "allowed_values", "aliases",
        "description", "confidence_field", "extraction_hints", "regex_patterns",
    ]
    ws.append(headers)

    # ── Guia Seção 2 (Template YAML) + Seções 3-9 (definições detalhadas) ──
    # NOTA: Cada campo, cada valor de enum e cada descrição abaixo foi
    #       verificado contra o texto do Guia de Extração v3.3.
    # fmt: off
    fields = [
        # ═══════════════════════ METADATA (Guia §2) ═══════════════════════
        ["ArtigoID", "Artigo ID", "metadata", "string", "true", "false",
         0, 0, "", "",
         "ID do artigo fornecido pelo solicitante. Formato: ART_0XX. "
         "NÃO inventar — usar exatamente o ID recebido.",
         "", "ART_001|ART_002", r"^ART_\d{3}$"],

        ["Ano", "Ano de Publicação", "metadata", "int", "true", "false",
         0, 0, "", "",
         "Ano de publicação do artigo (4 dígitos).",
         "", "", ""],

        ["TipoPublicação", "Tipo de Publicação", "metadata", "enum", "true", "false",
         0, 0,
         "Journal|Conference|Relatório técnico|Tese/Dissertação|Outro|NR",
         "TipoPublicacao",
         "Tipo do veículo de publicação.",
         "", "", ""],

        ["Referência_Curta", "Referência Curta", "metadata", "string", "true", "false",
         0, 0, "", "Referencia_Curta",
         "Formato: 'Autor et al., Ano' (ex: 'Silva et al., 2022').",
         "", "", ""],

        ["DOI", "DOI", "metadata", "string", "false", "false",
         0, 0, "", "",
         "Digital Object Identifier (https://doi.org/...).",
         "", "", ""],

        # ═══════════════════════ CONTEXTO — C (Guia §3) ═══════════════════
        ["SegmentoO&G", "Segmento O&G", "context", "enum", "true", "false",
         0, 0,
         "Upstream (E&P)|Midstream|Downstream|Cross-segment|NR",
         "SegmentoSetorial|SegmentoOG",
         "Segmento da indústria de O&G onde a intervenção de IA é aplicada. "
         "Upstream=exploração/perfuração/produção; Midstream=transporte/escoamento/armazenamento; "
         "Downstream=refino/processamento/distribuição; Cross-segment=múltiplos segmentos integrados.",
         "SegmentoO&G_Confiança",
         "upstream|midstream|downstream|exploration|refining|drilling|subsea|pipeline|refinery",
         ""],

        ["SegmentoO&G_Confiança", "Confiança Segmento", "context", "enum", "true", "false",
         0, 0, "Alta|Média|Baixa", "SegmentoSetorial_Confiança",
         "Nível de confiança na classificação do segmento.",
         "", "", ""],

        ["Ambiente", "Ambiente", "context", "enum", "true", "false",
         0, 0, "Offshore|Onshore|Híbrido|NR", "",
         "REGRA DE OURO: Classificar onde a INTERVENÇÃO DE IA é aplicada, "
         "NÃO onde o ativo final operará. Pergunte: (1) Onde os dados da IA são coletados/processados? "
         "(2) Onde a decisão apoiada pela IA é executada? "
         "ARMADILHA: Não confundir o PRODUTO final (ex: navio) com o PROCESSO onde a IA atua (ex: otimização de fabricação).",
         "",
         "offshore|onshore|field|platform|subsea|onshore facility|refinery|shipyard",
         ""],

        ["Complexidade", "Complexidade", "context", "enum", "true", "false",
         0, 0, "Alta|Média|Baixa|NR", "",
         "Sistema de pontuação OBRIGATÓRIO: "
         "F1=Ambiente offshore?(+1), F2=≥5 fornecedores mencionados?(+1), F3=Risco descrito como alto/crítico?(+1). "
         "Resultado: 3pts=Alta, 2pts=Média, 0-1pts=Baixa. "
         "Se fatores não avaliáveis → NR.",
         "", "", ""],

        ["Complexidade_Justificativa", "Justificativa Complexidade", "context", "string", "true", "false",
         10, 0, "", "",
         "DEVE conter pontuação explícita F1/F2/F3 com justificativa. "
         "Exemplo: 'F1=0 (onshore), F2=1 (>7000 fornecedores), F3=1 (risco crítico mencionado)'. "
         "NUNCA omitir os fatores.",
         "", "F1=|F2=|F3=", ""],

        ["ProcessoSCM_Alvo", "Processo SCM Alvo", "context", "enum", "true", "false",
         0, 0,
         "Planejamento de demanda|Gestão de estoques|Seleção/qualificação de fornecedores|"
         "Gestão de contratos|Procurement/compras|Logística/transporte|"
         "Planejamento de produção|Manutenção/MRO|Gestão de riscos da cadeia|"
         "Monitoramento/visibilidade|Integração/coordenação|Outro|NR",
         "",
         "Processo de SCM que a IA visa apoiar/otimizar.",
         "",
         "supply chain|supplier|procurement|logistics|inventory|maintenance|demand|production planning",
         ""],

        ["TipoRisco_SCRM", "Tipo Risco SCRM", "context", "enum", "true", "false",
         0, 0,
         "Risco de fornecimento|Risco de demanda|Risco operacional|"
         "Risco de transporte/logística|Risco de estoque|Risco financeiro|"
         "Risco de ativos|Risco ambiental/regulatório|Risco de segurança (security)|"
         "Risco de projeto|Risco geopolítico/externo|Múltiplos|NR",
         "",
         "Tipo do risco de supply chain endereçado pela IA.",
         "",
         "risk|disruption|failure|threat|vulnerability|supply risk|demand risk",
         ""],

        ["ObjetoCrítico", "Objeto Crítico", "context", "string", "false", "false",
         0, 0, "", "",
         "Ativo, recurso ou elemento em risco identificado no artigo. "
         "OBRIGATÓRIO se TipoRisco_SCRM ≠ NR (validado pela Regra R01).",
         "",
         "pipeline|equipment|supplier|asset|vessel|platform|compressor",
         ""],

        # ═══════════════════════ NARRATIVAS (Guia §8) ═════════════════════
        ["ProblemaNegócio_Contexto", "Problema/Contexto de Negócio", "narrative", "text", "true", "false",
         0, 0, "", "",
         "Problema de negócio + contexto operacional (3-6 linhas). "
         "Descrever O QUE motivou o estudo, em que contexto organizacional/industrial, "
         "e quais restrições/desafios existiam antes da intervenção de IA.",
         "",
         "problema|desafio|contexto|restrição|insuficiente|gap|challenge|current practice",
         ""],

        ["Intervenção_Descrição", "Descrição da Intervenção", "narrative", "text", "true", "false",
         0, 0, "", "",
         "Solução de IA implementada (2-5 linhas). Descrever: técnica usada, "
         "fluxo de processamento, variáveis de entrada/saída, e como a IA foi integrada ao processo. "
         "Se ClasseIA=Híbrido, listar TODAS as técnicas combinadas.",
         "",
         "modelo|algoritmo|técnica|framework|pipeline|approach|method|system",
         ""],

        ["Dados_Descrição", "Descrição dos Dados", "narrative", "text", "true", "false",
         0, 0, "", "",
         "Fontes, volume, período e qualidade dos dados (2-6 linhas). "
         "Incluir: origem dos dados, número de registros/atributos, período coberto, "
         "pré-processamento aplicado. NÃO confundir com Resultados_Quant.",
         "",
         "dataset|dados|registros|atributos|período|fonte|records|features|samples",
         ""],

        # ═══════════════ INTERVENÇÃO — I (Guia §4) ════════════════════════
        ["ClasseIA", "Classe de IA", "intervention", "enum", "true", "false",
         0, 0,
         "ML supervisionado|ML não supervisionado|Deep Learning|NLP|"
         "Visão computacional|Otimização/heurísticas|Probabilístico/Bayesiano|"
         "Sistemas especialistas/regras|Digital Twin com IA|Híbrido|Outro|NR",
         "",
         "Classe principal da técnica de IA. "
         "ML supervisionado=dados rotulados (XGBoost, RF, SVM). "
         "ML não supervisionado=clusters/anomalias sem rótulo (K-means, DBSCAN). "
         "Deep Learning=redes profundas (CNN, RNN, LSTM, Transformers, GAT). "
         "NLP=BERT, LLM, text mining. "
         "Se combina 2+ classes → 'Híbrido' (OBRIGATÓRIO detalhar em FamíliaModelo e Intervenção_Descrição).",
         "ClasseIA_Confiança",
         "machine learning|deep learning|neural|optimization|NLP|supervised|unsupervised",
         ""],

        ["ClasseIA_Confiança", "Confiança Classe IA", "intervention", "enum", "true", "false",
         0, 0, "Alta|Média|Baixa", "",
         "Nível de confiança na classificação da classe de IA.",
         "", "", ""],

        ["TarefaAnalítica", "Tarefa Analítica", "intervention", "enum", "true", "false",
         0, 0,
         "Previsão|Classificação|Regressão|Detecção de anomalia|"
         "Clustering/segmentação|Extração de informação (NLP)|Recomendação|"
         "Otimização|Simulação/what-if|Automação (RPA+IA)|Outro|NR",
         "",
         "Tarefa analítica que a IA executa. "
         "Previsão=forecasting de valores futuros. "
         "Classificação=atribuição a categorias discretas. "
         "Clustering/segmentação=agrupamento SEM rótulos.",
         "",
         "predict|classify|cluster|detect|optimize|recommend|forecast|segment",
         ""],

        ["FamíliaModelo", "Família de Modelo", "intervention", "enum", "true", "true",
         0, 0,
         "Ensemble tree-based|SVM/Kernel|Regressão linear/logística|"
         "Redes neurais feedforward|Redes recorrentes|Redes convolucionais|"
         "Transformers/Attention|Clustering|Redes Bayesianas|Metaheurísticas|"
         "Programação matemática|MCDM/Decisão|CBR/Especialista|Simulação|Outro|NR",
         "",
         "Algoritmo(s) específico(s). Se múltiplos, separar por ponto-e-vírgula. "
         "REGRAS DE CLASSIFICAÇÃO CRÍTICAS: "
         "(1) Ensemble=Random Forest, XGBoost, GBM, AdaBoost (múltiplas árvores agregadas). "
         "(2) Árvore de decisão ÚNICA (CHAID, CART, C4.5, ID3) → use 'Outro: Árvore de decisão'. "
         "(3) KNN/K-Nearest Neighbors (supervisionado) → use 'Outro: Instance-based/KNN' (NÃO é Clustering). "
         "(4) Clustering=K-means, DBSCAN, hierárquico (NÃO supervisionado, SEM labels). "
         "(5) Redes neurais feedforward=MLP, ANN, Perceptron.",
         "",
         "random forest|XGBoost|LSTM|CNN|BERT|AHP|SVM|GA|PSO|LightGBM|CatBoost|GRU|MLP|KNN",
         ""],

        ["TipoDado", "Tipo de Dado", "intervention", "enum", "true", "false",
         0, 0,
         "Tabular/ERP|Séries temporais|Texto|Imagem/Vídeo|Multimodal|NR",
         "",
         "Tipo principal dos dados usados pela IA.",
         "",
         "tabular|time series|text|image|multimodal|structured|ERP|sensor",
         ""],

        ["Maturidade", "Maturidade", "intervention", "enum", "true", "false",
         0, 0, "Conceito|Protótipo|Piloto|Produção|NR", "",
         "Nível de maturidade da implementação. "
         "Conceito=proposta teórica sem implementação. "
         "Protótipo=implementação experimental em laboratório. "
         "Piloto=teste em ambiente real com escopo limitado. "
         "Produção=sistema em operação contínua. "
         "REGRA: Se NívelEvidência='Simulação com dados sintéticos', Maturidade NÃO pode ser Produção nem Piloto.",
         "Maturidade_Confiança",
         "concept|prototype|pilot|production|case study|implementation|deployment",
         ""],

        ["Maturidade_Confiança", "Confiança Maturidade", "intervention", "enum", "true", "false",
         0, 0, "Alta|Média|Baixa", "",
         "Confiança na classificação de maturidade.",
         "", "", ""],

        # ═══════════════ MECANISMO — M (Guia §5) ══════════════════════════
        ["CategoriaMecanismo", "Categoria Mecanismo", "mechanism", "enum", "true", "false",
         0, 0,
         "Antecipação de risco|Detecção precoce/anomalias|Redução de incerteza|"
         "Priorização/alocação ótima|Integração de dados dispersos|"
         "Padronização/consistência|Otimização de trade-offs|"
         "Automação informacional (NLP)|Outro|NR",
         "",
         "CAMPO MAIS CRÍTICO: Como/por que a IA gera valor. "
         "Se o artigo não explica o porquê, use padrões de inferência: "
         "Prevê antes de ocorrer→Antecipação; Detecta anomalia→Detecção precoce; "
         "Melhora acurácia de forecast→Redução de incerteza; "
         "Processa múltiplas fontes→Integração de dados dispersos; "
         "Padroniza avaliação→Padronização/consistência; "
         "Otimiza múltiplos objetivos→Otimização de trade-offs.",
         "",
         "antecip|detect|reduz|prioriz|integr|padroniz|otimiz|automat|mechanism|why|how|value",
         ""],

        ["Mecanismo_Fonte", "Fonte do Mecanismo", "mechanism", "enum", "true", "false",
         0, 0, "Declarado|Inferido|Misto", "",
         "Se o mecanismo é declarado pelo autor, inferido pelo extrator, ou misto.",
         "", "", ""],

        ["Mecanismo_Declarado", "Mecanismo Declarado", "mechanism", "text", "false", "false",
         0, 0, "", "",
         "Transcrição de como o artigo EXPLICA o mecanismo. "
         "Usar NR se o artigo não declara explicitamente.",
         "",
         "mechanism|how|why|value|explains|contributes|enables|because",
         ""],

        ["Mecanismo_Inferido", "Mecanismo Inferido", "mechanism", "text", "false", "false",
         0, 0, "", "",
         "Inferência construída pelo extrator. "
         "REGRA ABSOLUTA: CADA sentença DEVE iniciar com o prefixo 'INFERIDO:'. "
         "Formato correto: 'INFERIDO: O clustering reduz espaço de busca. INFERIDO: A padronização melhora consistência.' "
         "Formato ERRADO: 'INFERIDO: O clustering reduz. A padronização...' (falta prefixo na 2ª sentença).",
         "",
         "INFERIDO:",
         ""],

        ["Mecanismo_Estruturado", "Mecanismo Estruturado", "mechanism", "string", "true", "false",
         10, 0, "", "",
         "String ÚNICA no formato: 'Entrada → Transformação → Mediação → Resultado'. "
         "Exemplo: 'Dados históricos → CBR + clustering → Recomendação → Seleção de fornecedor'. "
         "OBRIGATÓRIO: usar setas → (não hifens). NÃO usar múltiplas linhas ou bloco YAML multilinha.",
         "", "→", ""],

        # ═══════════════ OUTCOME — O (Guia §6) ════════════════════════════
        ["ResultadoTipo", "Tipo de Resultado", "outcome", "enum", "true", "false",
         0, 0, "Quantitativo|Qualitativo|Misto|NR", "",
         "Natureza dos resultados reportados no artigo.",
         "",
         "accuracy|precision|F1|improvement|qualitative|quantitative|performance",
         ""],

        ["Resultados_Quant", "Resultados Quantitativos", "outcome", "string", "false", "false",
         0, 0, "", "",
         "Formato OBRIGATÓRIO: 'métrica: valor (vs. baseline: X)' ou '(baseline: NR)' se não há baseline. "
         "Exemplo: 'Acurácia: 92.5% (vs. baseline: 78.3%); F1-score: 0.89 (baseline: NR)'. "
         "ARMADILHA: Tamanho de dataset ('1430 registros, 36 atributos') vai em Dados_Descrição, NÃO aqui.",
         "",
         "accuracy|error|precision|recall|F1|RMSE|MAE|improvement|reduction",
         ""],

        ["Resultados_Qual", "Resultados Qualitativos", "outcome", "text", "false", "false",
         0, 0, "", "",
         "Resultados descritivos/qualitativos reportados.",
         "", "", ""],

        ["NívelEvidência", "Nível de Evidência", "outcome", "enum", "true", "false",
         0, 0,
         "Estudo de caso real|Experimento com dados reais|"
         "Simulação com dados reais|Simulação com dados sintéticos|"
         "Survey/entrevistas|Proposta teórica/framework|Revisão conceitual|NR",
         "NivelEvidencia",
         "Robustez metodológica do estudo.",
         "",
         "case study|experiment|simulation|survey|theoretical|real data|synthetic",
         ""],

        ["Limitações_Artigo", "Limitações do Artigo", "outcome", "text", "true", "false",
         0, 0, "", "",
         "APENAS limitações RECONHECIDAS/DECLARADAS pelos autores do artigo. "
         "Se nenhuma limitação é declarada, usar NR — mas CONFIRMAR ausência buscando: "
         "limit*, restrict*, challeng*, only, future work, threats to validity.",
         "",
         "limitation|restriction|challenge|future work|threats to validity|weakness",
         ""],

        # ═══════════════ OPCIONAL ══════════════════════════════════════════
        ["Observação", "Observação", "optional", "text", "false", "false",
         0, 0, "", "",
         "Notas do extrator para casos ambíguos ou decisões de classificação não óbvias.",
         "", "", ""],
    ]
    # fmt: on
    for row in fields:
        ws.append(row)

    widths = {
        "A": 28, "B": 30, "C": 12, "D": 8, "E": 8, "F": 8,
        "G": 10, "H": 10, "I": 65, "J": 22, "K": 80, "L": 22, "M": 45, "N": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    _style_headers(ws, len(headers))


# ─────────────────────────────────────────────────────────────── rules
def _build_rules(wb: Workbook) -> None:
    """12 regras de validação do Guia §10 (Tabela de Regras Essenciais)."""
    ws = wb.create_sheet("rules")
    ws.append(["id", "severity", "when", "assert", "message"])

    # DSL disponível: get, eq, ne, not_empty, empty, contains, contains_any,
    #                 contains_all, regex, single_line, lower, in_set,
    #                 every_sentence_startswith
    # fmt: off
    rules = [
        # R01: TipoRisco ≠ NR → ObjetoCrítico obrigatório
        ["R01", "error",
         "ne(get('TipoRisco_SCRM'), 'NR')",
         "not_empty(get('ObjetoCrítico'))",
         "TipoRisco_SCRM ≠ NR → ObjetoCrítico DEVE estar preenchido"],

        # R02: Quantitativo → formato com baseline
        ["R02", "error",
         "eq(get('ResultadoTipo'), 'Quantitativo')",
         "contains(get('Resultados_Quant'), 'baseline')",
         "ResultadoTipo = Quantitativo → Resultados_Quant deve conter 'baseline' no formato 'métrica: valor (vs. baseline: X)'"],

        # R03: Produção → evidência robusta
        ["R03", "error",
         "eq(get('Maturidade'), 'Produção')",
         "in_set(get('NívelEvidência'), ['Estudo de caso real', 'Experimento com dados reais'])",
         "Maturidade = Produção → NívelEvidência deve ser 'Estudo de caso real' ou 'Experimento com dados reais'"],

        # R04: Híbrido → detalhar famílias
        ["R04", "warning",
         "eq(get('ClasseIA'), 'Híbrido')",
         "not_empty(get('FamíliaModelo'))",
         "ClasseIA = Híbrido → FamíliaModelo DEVE especificar as técnicas combinadas. Também detalhar em Intervenção_Descrição"],

        # R05: INFERIDO: em cada sentença
        ["R05", "error",
         "not_empty(get('Mecanismo_Inferido'))",
         "every_sentence_startswith(get('Mecanismo_Inferido'), 'INFERIDO:')",
         "Mecanismo_Inferido → CADA sentença deve iniciar com 'INFERIDO:'. Verifique TODAS as sentenças"],

        # R06: Mecanismo_Estruturado com setas
        ["R06", "error",
         "not_empty(get('Mecanismo_Estruturado'))",
         "contains(get('Mecanismo_Estruturado'), '→') and single_line(get('Mecanismo_Estruturado'))",
         "Mecanismo_Estruturado deve ser STRING ÚNICA com formato 'Entrada → Transformação → Mediação → Resultado'"],

        # R07: formato ArtigoID
        ["R07", "error",
         "True",
         "regex(get('ArtigoID'), '^ART_[0-9]{3}$')",
         "ArtigoID deve corresponder ao formato ART_0XX (ex: ART_001)"],

        # R08: simulação sintética → não pode ser Produção/Piloto
        ["R08", "error",
         "eq(get('NívelEvidência'), 'Simulação com dados sintéticos')",
         "not_empty(get('Maturidade')) and ne(get('Maturidade'), 'Produção') and ne(get('Maturidade'), 'Piloto')",
         "Simulação com dados sintéticos → Maturidade NÃO pode ser Produção nem Piloto"],

        # R09: F1/F2/F3 na justificativa
        ["R09", "error",
         "True",
         "regex(get('Complexidade_Justificativa'), 'F[123]=')",
         "Complexidade_Justificativa DEVE conter pontuação F1/F2/F3 explícita"],

        # R10: Ambiente preenchido (lembrete de regra de ouro)
        ["R10", "warning",
         "True",
         "not_empty(get('Ambiente'))",
         "Ambiente deve refletir onde a IA ATUA, NÃO onde o ativo final opera. Rever se necessário"],

        # R11: Limitações = NR → confirmar
        ["R11", "warning",
         "eq(get('Limitações_Artigo'), 'NR')",
         "True",
         "Limitações_Artigo = NR → Confirmar ausência de termos: limit*, restrict*, challeng*, only, future work, threats to validity"],

        # R12: ArtigoID obrigatório (proxy para campos exatos)
        ["R12", "error",
         "True",
         "not_empty(get('ArtigoID'))",
         "Campos obrigatórios devem usar nomes EXATOS do template (sem variações)"],
    ]
    # fmt: on
    for row in rules:
        ws.append(row)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 85
    ws.column_dimensions["E"].width = 100
    _style_headers(ws, 5)


# ─────────────────────────────────────────────────────────────── quotes
def _build_quotes_policy(wb: Workbook) -> None:
    """Guia §7 — Diretrizes de Quotes."""
    ws = wb.create_sheet("quotes_policy")
    ws.append(["key", "value"])
    rows = [
        ("enabled", "true"),
        ("min_quotes", "3"),
        ("max_quotes", "8"),
        ("required_types", "Mecanismo"),
        ("allowed_types", "Contexto|Intervenção|Mecanismo|Outcome|Limitação|Método|Outro"),
        ("id_pattern", r"^Q\d{3}$"),
        ("required_fields", "QuoteID|TipoQuote|Trecho|Página"),
        ("trecho_min_length", "10"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 65
    _style_headers(ws, 2)


# ─────────────────────────────────────────────────────────────── prompt
def _build_prompt(wb: Workbook) -> None:
    """Instruções condensadas do Guia v3.3 para o LLM (complementa guia_v3_3_prompt.md)."""
    ws = wb.create_sheet("prompt")
    ws.append(["instruction"])
    instructions = [
        # Papel e framework
        "Você é o extrator de dados de uma RSL sobre IA + SCM/SCRM em Oil & Gas, seguindo o framework CIMO (Context-Intervention-Mechanism-Outcome).",
        "Retorne SOMENTE YAML válido conforme o template do perfil ativo. Não invente valores; quando não houver evidência explícita, use NR.",
        # Mecanismo — campo mais crítico
        "PRIORIDADE ABSOLUTA: Mecanismo (M) é o campo MAIS CRÍTICO. Identifique HOW/WHY a IA gera valor. Se o artigo não explica, INFIRA usando padrões: prevê→Antecipação; detecta anomalia→Detecção precoce; melhora forecast→Redução de incerteza; processa múltiplas fontes→Integração; padroniza→Padronização; otimiza→Otimização de trade-offs.",
        "CADA sentença em Mecanismo_Inferido DEVE iniciar com 'INFERIDO:'. Formato errado: 'INFERIDO: A reduz. B melhora.' Formato correto: 'INFERIDO: A reduz. INFERIDO: B melhora.'",
        "Mecanismo_Estruturado: STRING ÚNICA no formato 'Entrada → Transformação → Mediação → Resultado'. NÃO usar múltiplas linhas.",
        # Preenchimento crítico
        "Resultados_Quant: formato 'métrica: valor (vs. baseline: X)' ou '(baseline: NR)'. NUNCA incluir tamanho de dataset aqui — isso vai em Dados_Descrição.",
        "Complexidade_Justificativa DEVE conter pontuação F1/F2/F3 explícita com justificativa para cada fator.",
        "Ambiente: REGRA DE OURO — classificar onde a IA ATUA, não onde o ativo final opera. Fabricação em estaleiro para FPSO = Onshore.",
        "Limitações_Artigo: APENAS reconhecidas pelos autores. Se NR, confirmar buscando: limit*, restrict*, challeng*, threats to validity.",
        # Classificação de modelos
        "FamíliaModelo — REGRAS: Ensemble=RF/XGBoost/GBM/AdaBoost; Árvore única (CHAID/CART)→'Outro: Árvore de decisão'; KNN→'Outro: Instance-based/KNN' (NÃO é Clustering); Clustering=K-means/DBSCAN (não supervisionado, sem labels); Feedforward=MLP/ANN/Perceptron.",
        "Se ClasseIA = 'Híbrido', OBRIGATÓRIO listar técnicas em FamíliaModelo (separando por ;) E detalhar em Intervenção_Descrição.",
        # Quotes
        "Quotes: 3-8 quotes LITERAIS (cópia exata do texto), ≤3 linhas cada, com Página. Priorizar pelo menos 1 quote de Mecanismo. Formato: QuoteID (Q001), TipoQuote, Trecho, Página (p.X).",
        # Auto-revisão
        "PROTOCOLO DE AUTO-REVISÃO (executar antes de finalizar): FASE 1 (maior risco): Ambiente reflete IA? Complexidade tem F1/F2/F3? Mecanismo_Estruturado string única? INFERIDO: em cada sentença? FASE 2: Híbrido detalhado? Maturidade consistente com evidência? FASE 3: Campos obrigatórios? 3-8 quotes? FASE 4: 12 regras verificadas?",
    ]
    for inst in instructions:
        ws.append([inst])
    ws.column_dimensions["A"].width = 140
    _style_headers(ws, 1)


# ─────────────────────────────────────────────────────────────── README
def _build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    ws["A1"] = "SAEC Profile Template — RSL IA + SCM/SCRM em Oil & Gas (CIMO v3.3)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Este XLSX configura o perfil de extração do SAEC para a pesquisa de mestrado."
    ws["A4"] = "Derivado fielmente do Guia de Extração v3.3 (Guia_Extracao_v3_3.docx/md)."
    ws["A5"] = ""
    ws["A6"] = "Abas:"
    ws["A7"] = "  • meta — Metadados do perfil (ID, versão, framework CIMO, domínio O&G SCM)"
    ws["A8"] = "  • fields — 34 campos CIMO com tipos, enums exatos do Guia, regras de classificação e hints de extração"
    ws["A9"] = "  • rules — 12 regras de validação obrigatórias (exatas do Guia §10)"
    ws["A10"] = "  • quotes_policy — Política de quotes: 3-8, tipos obrigatórios (Mecanismo), schema QuoteID/TipoQuote/Trecho/Página"
    ws["A11"] = "  • prompt — Instruções condensadas para o LLM (complementa config/prompts/guia_v3_3_prompt.md)"
    ws["A12"] = ""
    ws["A13"] = "Para usar: GUI → Profile → Import XLSX → selecione este arquivo."
    ws["A14"] = "O prompt completo (254 linhas) está em: config/prompts/guia_v3_3_prompt.md"
    ws.column_dimensions["A"].width = 100


def main() -> None:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _build_readme(wb)
    _build_meta(wb)
    _build_fields(wb)
    _build_rules(wb)
    _build_quotes_policy(wb)
    _build_prompt(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"✅ Profile XLSX gerado: {OUTPUT}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Fields: {sum(1 for _ in wb['fields'].iter_rows(min_row=2)) if 'fields' in wb.sheetnames else 0}")
    print(f"   Rules:  {sum(1 for _ in wb['rules'].iter_rows(min_row=2)) if 'rules' in wb.sheetnames else 0}")
    print(f"   Prompt: {sum(1 for _ in wb['prompt'].iter_rows(min_row=2)) if 'prompt' in wb.sheetnames else 0} instruções")


if __name__ == "__main__":
    main()
