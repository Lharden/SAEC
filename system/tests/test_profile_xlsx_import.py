from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from profile_engine.loader import ProfileLoadError
from profile_engine.project_profiles import (
    get_active_profile_ref,
    import_profile_xlsx,
    load_active_profile_spec,
)
from profile_engine.xlsx_profiles import (
    load_profile_spec_from_xlsx,
    write_profile_template_xlsx,
)


def _init_project(tmp_path: Path) -> Path:
    project_root = tmp_path / "project"
    (project_root / "config" / "profiles").mkdir(parents=True, exist_ok=True)
    manifest = {
        "project_id": "tmp",
        "name": "tmp",
        "created_at": "2026-02-19T00:00:00+00:00",
        "profile_required": True,
    }
    (project_root / "project.json").write_text(
        json.dumps(manifest),
        encoding="utf-8",
    )
    return project_root


def test_write_template_and_load_profile_spec_from_xlsx(tmp_path: Path) -> None:
    template = tmp_path / "profile_template.xlsx"
    write_profile_template_xlsx(template)

    result = load_profile_spec_from_xlsx(template)

    assert result.spec.meta.profile_id == "custom_profile"
    assert result.spec.field_by_id("ArtigoID") is not None
    assert "TemaPrincipal" in result.prompt_text
    assert "termos-chave" in result.prompt_text


def test_import_profile_xlsx_activates_project_profile(tmp_path: Path) -> None:
    project_root = _init_project(tmp_path)
    template = tmp_path / "profile_template.xlsx"
    write_profile_template_xlsx(template)

    ref = import_profile_xlsx(project_root, xlsx_path=template, activate=True)
    active = get_active_profile_ref(project_root)
    spec, _ = load_active_profile_spec(project_root)

    assert ref.source == "xlsx_import"
    assert active is not None
    assert active.profile_id == spec.meta.profile_id
    assert spec.field_by_id("ArtigoID") is not None


def test_load_profile_spec_from_xlsx_requires_all_sheets(tmp_path: Path) -> None:
    broken = tmp_path / "broken.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "meta"
    ws.append(["key", "value"])
    ws.append(["profile_id", "broken"])
    ws.append(["version", "1.0.0"])
    ws.append(["name", "Broken"])
    wb.save(broken)

    with pytest.raises(ProfileLoadError):
        load_profile_spec_from_xlsx(broken)


def test_load_profile_spec_from_xlsx_reports_row_errors(tmp_path: Path) -> None:
    template = tmp_path / "profile_template.xlsx"
    write_profile_template_xlsx(template)
    wb = load_workbook(template)
    ws = wb["fields"]
    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    required_col = headers.index("required") + 1
    ws.cell(row=2, column=required_col).value = "maybe"
    wb.save(template)

    with pytest.raises(ProfileLoadError) as exc_info:
        load_profile_spec_from_xlsx(template)
    assert "fields row 2" in str(exc_info.value)


def test_load_profile_spec_from_xlsx_rejects_invalid_rule_expr(tmp_path: Path) -> None:
    template = tmp_path / "profile_template.xlsx"
    write_profile_template_xlsx(template)
    wb = load_workbook(template)
    ws = wb["rules"]
    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    assert_col = headers.index("assert") + 1
    ws.cell(row=2, column=assert_col).value = "unknown_func(get('ArtigoID'))"
    wb.save(template)

    with pytest.raises(ProfileLoadError):
        load_profile_spec_from_xlsx(template)
