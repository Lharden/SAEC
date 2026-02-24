"""Import/export helpers for XLSX-based profile templates."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .loader import ProfileLoadError
from .models import CURRENT_PROFILE_SCHEMA_VERSION, ProfileSpec
from .validator import validate_rule_expressions


_SHEET_META = "meta"
_SHEET_FIELDS = "fields"
_SHEET_RULES = "rules"
_SHEET_QUOTES = "quotes_policy"
_SHEET_PROMPT = "prompt"
_SHEET_README = "README"

_FIELD_HEADERS = [
    "id",
    "label",
    "section",
    "type",
    "required",
    "multiple",
    "min_length",
    "max_length",
    "allowed_values",
    "aliases",
    "description",
    "confidence_field",
    "extraction_hints",
    "regex_patterns",
]
_RULE_HEADERS = ["id", "severity", "when", "assert", "message"]
_META_HEADERS = ["key", "value"]
_PROMPT_HEADERS = ["instruction"]


@dataclass(frozen=True)
class XlsxProfileImportResult:
    spec: ProfileSpec
    prompt_text: str
    notes: tuple[str, ...]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _cell_text(value).lower()


def _list_from_cell(value: Any) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    if "|" in text:
        tokens = text.split("|")
    else:
        tokens = text.split(",")
    return [token.strip() for token in tokens if token.strip()]


def _to_bool(value: Any, *, default: bool) -> bool:
    text = _cell_text(value).lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "on", "sim", "s"}:
        return True
    if text in {"0", "false", "no", "n", "off", "nao", "não"}:
        return False
    raise ValueError(f"invalid boolean value '{value}'")


def _to_int(value: Any, *, default: int) -> int:
    text = _cell_text(value)
    if not text:
        return default
    return int(float(text))


def _resolve_headers(ws, expected_headers: list[str], *, sheet_name: str) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        header = _normalize_header(cell.value)
        if header:
            header_map[header] = idx
    missing = [name for name in expected_headers if name not in header_map]
    if missing:
        raise ProfileLoadError(
            f"XLSX sheet '{sheet_name}' is missing required header(s)",
            errors=tuple(missing),
        )
    return header_map


def _iter_rows(ws, *, min_row: int = 2):
    for row_idx in range(min_row, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        if not any(_cell_text(value) for value in values):
            continue
        yield row_idx, values


def _parse_meta(ws) -> dict[str, str]:
    headers = _resolve_headers(ws, _META_HEADERS, sheet_name=_SHEET_META)
    payload: dict[str, str] = {}
    for row_idx, _ in _iter_rows(ws):
        key = _cell_text(ws.cell(row=row_idx, column=headers["key"]).value)
        value = _cell_text(ws.cell(row=row_idx, column=headers["value"]).value)
        if not key:
            continue
        payload[key] = value
    return payload


def _parse_fields(ws) -> tuple[list[dict[str, Any]], list[str]]:
    headers = _resolve_headers(ws, _FIELD_HEADERS, sheet_name=_SHEET_FIELDS)
    fields: list[dict[str, Any]] = []
    errors: list[str] = []
    seen_ids: set[str] = set()
    for row_idx, _ in _iter_rows(ws):
        field_id = _cell_text(ws.cell(row=row_idx, column=headers["id"]).value)
        if not field_id:
            errors.append(f"{_SHEET_FIELDS} row {row_idx}: id is required")
            continue
        if field_id in seen_ids:
            errors.append(f"{_SHEET_FIELDS} row {row_idx}: duplicate id '{field_id}'")
            continue
        seen_ids.add(field_id)
        field_type = _cell_text(ws.cell(row=row_idx, column=headers["type"]).value).lower() or "string"
        try:
            required = _to_bool(
                ws.cell(row=row_idx, column=headers["required"]).value,
                default=True,
            )
            multiple = _to_bool(
                ws.cell(row=row_idx, column=headers["multiple"]).value,
                default=False,
            )
            min_length = _to_int(
                ws.cell(row=row_idx, column=headers["min_length"]).value,
                default=0,
            )
            max_length = _to_int(
                ws.cell(row=row_idx, column=headers["max_length"]).value,
                default=0,
            )
        except ValueError as exc:
            errors.append(f"{_SHEET_FIELDS} row {row_idx}: {exc}")
            continue
        fields.append(
            {
                "id": field_id,
                "label": _cell_text(ws.cell(row=row_idx, column=headers["label"]).value) or field_id,
                "section": _cell_text(ws.cell(row=row_idx, column=headers["section"]).value) or "custom",
                "type": field_type,
                "required": required,
                "multiple": multiple,
                "min_length": max(min_length, 0),
                "max_length": max(max_length, 0),
                "allowed_values": _list_from_cell(
                    ws.cell(row=row_idx, column=headers["allowed_values"]).value
                ),
                "aliases": _list_from_cell(ws.cell(row=row_idx, column=headers["aliases"]).value),
                "description": _cell_text(
                    ws.cell(row=row_idx, column=headers["description"]).value
                ),
                "confidence_field": _cell_text(
                    ws.cell(row=row_idx, column=headers["confidence_field"]).value
                ),
                "extraction_hints": _list_from_cell(
                    ws.cell(row=row_idx, column=headers["extraction_hints"]).value
                ),
                "regex_patterns": _list_from_cell(
                    ws.cell(row=row_idx, column=headers["regex_patterns"]).value
                ),
            }
        )
    return fields, errors


def _parse_rules(ws) -> tuple[list[dict[str, str]], list[str]]:
    headers = _resolve_headers(ws, _RULE_HEADERS, sheet_name=_SHEET_RULES)
    rules: list[dict[str, str]] = []
    errors: list[str] = []
    seen_ids: set[str] = set()
    for row_idx, _ in _iter_rows(ws):
        rule_id = _cell_text(ws.cell(row=row_idx, column=headers["id"]).value)
        when_expr = _cell_text(ws.cell(row=row_idx, column=headers["when"]).value)
        assert_expr = _cell_text(ws.cell(row=row_idx, column=headers["assert"]).value)
        message = _cell_text(ws.cell(row=row_idx, column=headers["message"]).value)
        if not rule_id and not when_expr and not assert_expr and not message:
            continue
        if not rule_id:
            errors.append(f"{_SHEET_RULES} row {row_idx}: id is required")
            continue
        if rule_id in seen_ids:
            errors.append(f"{_SHEET_RULES} row {row_idx}: duplicate id '{rule_id}'")
            continue
        seen_ids.add(rule_id)
        severity = _cell_text(ws.cell(row=row_idx, column=headers["severity"]).value).lower() or "error"
        rules.append(
            {
                "id": rule_id,
                "severity": severity,
                "when": when_expr or "True",
                "assert": assert_expr or "True",
                "message": message or "Rule failed",
            }
        )
    return rules, errors


def _parse_quotes_policy(ws) -> tuple[dict[str, Any], list[str]]:
    headers = _resolve_headers(ws, _META_HEADERS, sheet_name=_SHEET_QUOTES)
    values: dict[str, str] = {}
    errors: list[str] = []
    for row_idx, _ in _iter_rows(ws):
        key = _cell_text(ws.cell(row=row_idx, column=headers["key"]).value)
        value = _cell_text(ws.cell(row=row_idx, column=headers["value"]).value)
        if not key:
            continue
        values[key] = value

    try:
        enabled = _to_bool(values.get("enabled"), default=True)
        min_quotes = _to_int(values.get("min_quotes"), default=3)
        max_quotes = _to_int(values.get("max_quotes"), default=8)
        trecho_min_length = _to_int(values.get("trecho_min_length"), default=10)
    except ValueError as exc:
        errors.append(f"{_SHEET_QUOTES}: {exc}")
        enabled = True
        min_quotes = 3
        max_quotes = 8
        trecho_min_length = 10

    quotes_policy = {
        "enabled": enabled,
        "min_quotes": max(min_quotes, 0),
        "max_quotes": max(max_quotes, 0),
        "required_types": _list_from_cell(values.get("required_types")),
        "allowed_types": _list_from_cell(values.get("allowed_types")),
        "quote_schema": {
            "id_pattern": values.get("id_pattern") or r"^Q\d{3}$",
            "required_fields": _list_from_cell(values.get("required_fields"))
            or ["QuoteID", "TipoQuote", "Trecho", "Página"],
            "trecho_min_length": max(trecho_min_length, 1),
        },
    }
    return quotes_policy, errors


def _parse_prompt_instructions(ws) -> list[str]:
    headers = _resolve_headers(ws, _PROMPT_HEADERS, sheet_name=_SHEET_PROMPT)
    instructions: list[str] = []
    for row_idx, _ in _iter_rows(ws):
        instruction = _cell_text(ws.cell(row=row_idx, column=headers["instruction"]).value)
        if instruction:
            instructions.append(instruction)
    return instructions


def _build_prompt_text(spec: ProfileSpec, prompt_instructions: list[str]) -> str:
    lines: list[str] = [
        f"# Extract Prompt ({spec.meta.framework or 'CUSTOM'})",
        "",
        "Retorne somente YAML válido conforme o perfil ativo.",
        "Não invente valores; quando não houver evidência explícita, use NR/empty conforme o campo.",
        "Cada preenchimento deve ser ancorado no texto do artigo.",
        "",
        "Campos esperados:",
    ]
    for field in spec.fields:
        required = "obrigatório" if field.required else "opcional"
        pieces: list[str] = [f"- {field.field_id} ({field.field_type}, {required})"]
        if field.allowed_values:
            pieces.append("enum: " + ", ".join(field.allowed_values))
        if field.extraction_hints:
            pieces.append("termos-chave: " + ", ".join(field.extraction_hints))
        if field.regex_patterns:
            pieces.append("regex: " + "; ".join(field.regex_patterns))
        if field.description:
            pieces.append("descrição: " + field.description)
        lines.append(" | ".join(pieces))

    if spec.rules:
        lines.extend(["", "Regras de consistência obrigatórias:"])
        for rule in spec.rules:
            lines.append(f"- {rule.rule_id}: {rule.message}")

    if spec.quotes_policy.enabled:
        lines.extend(
            [
                "",
                "Política de quotes:",
                f"- quantidade: {spec.quotes_policy.min_quotes} a {spec.quotes_policy.max_quotes}",
                "- schema quote: "
                + ", ".join(spec.quotes_policy.quote_schema.required_fields),
            ]
        )

    if prompt_instructions:
        lines.extend(["", "Instruções adicionais:"])
        for instruction in prompt_instructions:
            lines.append(f"- {instruction}")

    lines.append("")
    return "\n".join(lines)


def load_profile_spec_from_xlsx(xlsx_path: Path) -> XlsxProfileImportResult:
    """Parse one XLSX template and convert it to a validated ProfileSpec."""
    path = Path(xlsx_path).resolve()
    if not path.exists():
        raise ProfileLoadError(f"XLSX profile template not found: {path}")

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ProfileLoadError(
            f"Failed to open XLSX profile template: {path}",
            errors=(str(exc),),
        ) from exc

    required_sheets = {_SHEET_META, _SHEET_FIELDS, _SHEET_RULES, _SHEET_QUOTES, _SHEET_PROMPT}
    missing_sheets = sorted(name for name in required_sheets if name not in workbook.sheetnames)
    if missing_sheets:
        raise ProfileLoadError(
            "XLSX profile template is missing required sheet(s)",
            errors=tuple(missing_sheets),
        )

    meta = _parse_meta(workbook[_SHEET_META])
    fields, field_errors = _parse_fields(workbook[_SHEET_FIELDS])
    rules, rule_errors = _parse_rules(workbook[_SHEET_RULES])
    quotes_policy, quote_errors = _parse_quotes_policy(workbook[_SHEET_QUOTES])
    prompt_instructions = _parse_prompt_instructions(workbook[_SHEET_PROMPT])

    parse_errors = field_errors + rule_errors + quote_errors
    if parse_errors:
        raise ProfileLoadError(
            "XLSX profile template contains invalid rows",
            errors=tuple(parse_errors),
        )

    profile_payload = {
        "schema_version": meta.get("schema_version", CURRENT_PROFILE_SCHEMA_VERSION)
        or CURRENT_PROFILE_SCHEMA_VERSION,
        "profile": {
            "meta": {
                "profile_id": meta.get("profile_id", ""),
                "version": meta.get("version", ""),
                "name": meta.get("name", ""),
                "framework": meta.get("framework", "CUSTOM"),
                "language": meta.get("language", "pt-BR"),
                "domain": meta.get("domain", ""),
                "description": meta.get("description", ""),
            },
            "topics": {
                "min_topics": _to_int(meta.get("min_topics"), default=1),
                "max_topics": _to_int(meta.get("max_topics"), default=50),
                "weighting_mode": meta.get("weighting_mode", "optional") or "optional",
                "default_weight": float(meta.get("default_weight") or 1.0),
            },
            "fields": fields,
            "rules": rules,
            "quotes_policy": quotes_policy,
            "output": {
                "format": meta.get("output_format", "yaml") or "yaml",
                "enforce_field_names": _to_bool(
                    meta.get("enforce_field_names"),
                    default=True,
                ),
                "include_sections": _to_bool(
                    meta.get("include_sections"),
                    default=True,
                ),
            },
            "prompt_contract": {
                "return_only_output": _to_bool(
                    meta.get("return_only_output"),
                    default=True,
                ),
                "include_self_review": _to_bool(
                    meta.get("include_self_review"),
                    default=True,
                ),
                "instructions": prompt_instructions,
            },
        },
    }

    spec = ProfileSpec.from_dict(profile_payload)
    structure_errors = spec.validate_structure()
    structure_errors.extend(validate_rule_expressions(spec.rules))
    if structure_errors:
        raise ProfileLoadError(
            "XLSX template could not be converted to a valid profile",
            errors=tuple(structure_errors),
        )

    prompt_text = _build_prompt_text(spec, prompt_instructions)
    notes = (
        "xlsx_import",
        f"fields={len(spec.fields)}",
        f"rules={len(spec.rules)}",
        f"framework={spec.meta.framework}",
    )
    return XlsxProfileImportResult(spec=spec, prompt_text=prompt_text, notes=notes)


def write_profile_template_xlsx(destination: Path) -> Path:
    """Write an official editable XLSX template for profile configuration."""
    target = Path(destination).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_readme = wb.active
    if ws_readme is None:  # defensive for static typing; openpyxl always provides active sheet
        ws_readme = wb.create_sheet(_SHEET_README)
    ws_readme.title = _SHEET_README
    ws_readme["A1"] = "Profile Template XLSX (official)"
    ws_readme["A2"] = "Preencha as abas meta, fields, rules, quotes_policy e prompt."
    ws_readme["A3"] = "Use '|' para listas (ex.: enum1|enum2|enum3)."
    ws_readme["A4"] = "Campos obrigatórios mínimos: profile_id, version, name e field ArtigoID."
    ws_readme["A5"] = "Valores booleanos aceitos: true/false, yes/no, sim/nao, 1/0."

    ws_meta = wb.create_sheet(_SHEET_META)
    ws_meta.append(_META_HEADERS)
    meta_rows = [
        ("schema_version", CURRENT_PROFILE_SCHEMA_VERSION),
        ("profile_id", "custom_profile"),
        ("version", "1.0.0"),
        ("name", "Custom RSL Profile"),
        ("framework", "CUSTOM"),
        ("language", "pt-BR"),
        ("domain", ""),
        ("description", ""),
        ("min_topics", "1"),
        ("max_topics", "50"),
        ("weighting_mode", "optional"),
        ("default_weight", "1.0"),
        ("output_format", "yaml"),
        ("enforce_field_names", "true"),
        ("include_sections", "true"),
        ("return_only_output", "true"),
        ("include_self_review", "true"),
    ]
    for row in meta_rows:
        ws_meta.append(row)

    ws_fields = wb.create_sheet(_SHEET_FIELDS)
    ws_fields.append(_FIELD_HEADERS)
    ws_fields.append(
        [
            "ArtigoID",
            "ArtigoID",
            "metadata",
            "string",
            "true",
            "false",
            "0",
            "0",
            "",
            "",
            "ID do artigo no mapping",
            "",
            "ART_001|ART_002",
            "^ART_[0-9]{3}$",
        ]
    )
    ws_fields.append(
        [
            "TemaPrincipal",
            "TemaPrincipal",
            "custom",
            "enum",
            "true",
            "false",
            "0",
            "0",
            "A|B|C|NR",
            "",
            "Tema principal do estudo",
            "",
            "tema principal|problema central|objetivo",
            "",
        ]
    )

    ws_rules = wb.create_sheet(_SHEET_RULES)
    ws_rules.append(_RULE_HEADERS)
    ws_rules.append(
        [
            "R1",
            "error",
            "True",
            "regex(get('ArtigoID'), '^ART_[0-9]{3}$')",
            "ArtigoID deve estar no formato ART_001.",
        ]
    )

    ws_quotes = wb.create_sheet(_SHEET_QUOTES)
    ws_quotes.append(_META_HEADERS)
    quote_rows = [
        ("enabled", "true"),
        ("min_quotes", "3"),
        ("max_quotes", "8"),
        ("required_types", "Mecanismo"),
        ("allowed_types", "Contexto|Intervenção|Mecanismo|Outcome|Outro"),
        ("id_pattern", r"^Q\d{3}$"),
        ("required_fields", "QuoteID|TipoQuote|Trecho|Página"),
        ("trecho_min_length", "10"),
    ]
    for row in quote_rows:
        ws_quotes.append(row)

    ws_prompt = wb.create_sheet(_SHEET_PROMPT)
    ws_prompt.append(_PROMPT_HEADERS)
    ws_prompt.append(("Retorne somente YAML.",))
    ws_prompt.append(("Use apenas campos declarados no perfil.",))
    ws_prompt.append(("Não invente valores sem evidência textual.",))

    wb.save(target)
    return target
